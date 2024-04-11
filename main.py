"""A kanban board is used to track the status and priorities of various subtasks to be done in a software project.
Implement the following features:
Implement a basic kanban board to be used from a command line interface with 3 statuses (Todo, In progress, and Done)
Addition and removal of tasks to each, and moving them between statuses
Assignees and Reporters
Add more status fields to the Kanban Board
Support for multiple boards for different projects
Allow priorities and sorting by priority (High, Medium, Low)
Brownie points:

Usage of a mouse within the CLI to have drag and drop functionality.

Please make use of an appropriate storage solution to persist tasks, boards, and statuses across runs"""

from openpyxl import *
import pandas as pd
status = ['Todo','In Progress','Done']
def new_project():
    wb = load_workbook('kanban_board.xlsx')
    name = input("Enter the name of the project:")
    ws = wb.create_sheet(name)
    fields = ('id','title','status','assignee','reporter','priority')
    ws.append(fields)
    wb.save('kanban_board.xlsx')

def display_project():
    wb = load_workbook('kanban_board.xlsx')
    for i in wb.sheetnames:
        print(i)

'''def save_tasks():
    wb = load_workbook('kanban_board.xlsx')
    wb.save('kanban_board.xlsx')'''

def display_tasks(project):
    wb = load_workbook('kanban_board.xlsx')
    ws = wb[project]
    for task in ws.iter_rows(values_only=True):
        if task[1] == 'title':
            continue
        print(task[0],task[1],task[2],task[5])
    #(1)id;(2)title;(3)status;(4)assignee;(5)reporter;(6)priority

def add_task(project):
    wb = load_workbook('kanban_board.xlsx')
    ws = wb[project]
    title = input("Enter task title: ")
    assignee = input("Enter assignee: ")
    reporter = input("Enter reporter: ")
    priority = input("Enter priority (High, Medium, Low): ")
    task = (ws.max_row + 1,title,'Todo',assignee,reporter,priority)
    ws.append(task)
    wb.save('kanban_board.xlsx')
    print("Task added successfully.")

def move_task(project):
    wb = load_workbook('kanban_board.xlsx')
    ws = wb[project]
    display_tasks(project)
    task_id = int(input("Enter task id to move: "))
    for i in range(len(status)):
        print(i, status[i])
    new_status = input("Enter new status: ")
    for task in ws.iter_rows(values_only=True):
        if int(task[0]) == task_id:              #(1)id;(2)title;(3)status;(4)assignee;(5)reporter;(6)priority
            task[2] = status[new_status]
            break
    wb.save('kanban_board.xlsx')
    print("Task moved successfully")

def remove_task(project):
    wb = load_workbook('kanban_board.xlsx')
    ws = wb[project]
    display_tasks(project)
    task_id = int(input("Enter task id to remove: "))
    a = 2
    for task in ws.iter_rows(values_only=True):
        if task[0] == task_id:                      #(1)id;(2)title;(3)status;(4)assignee;(5)reporter;(6)priority
            ws.delete_rows(task_id)
    wb.save('kanban_board.xlsx')
    print("Task deleted successfully")

def edit_task(project):
    wb = load_workbook('kanban_board.xlsx')
    ws = wb[project]
    display_tasks(project)
    task_id = int(input("Enter task id to edit: "))
    for i in ws.iter_rows(values_only=True):                    #(1)id;(2)title;(3)status;(4)assignee;(5)reporter;(6)priority
        if i[0] == task_id:
            print(f"{i[0]}: {i[1]} {i[2]} {i[5]}\n Assignee:{i[3]}\n Reporter:{i[4]}")
    choice = 0
    while choice!=5:
        print("Edit options\n1.Title\n2.Priority\n3.Assignee\n4.Reporter\n5.Exit")
        choice = int(input("Enter the option:"))
        if choice == 1:
            new_title = input("Enter new title: ")
            for i in ws.iter_rows(values_only=True):
                if i[0] == task_id:
                    i[1] = new_title
        elif choice == 2:
            new_priority = input("Enter priority (High, Medium, Low): ")
            for i in ws.iter_rows(values_only=True):
                if i[0] == task_id:
                    i[5] = new_priority
        elif choice == 3:
            new_assignee = input("Enter assignee: ")
            for i in ws.iter_rows(values_only=True):
                if i[0] == task_id:
                    i[3] = new_assignee
        elif choice == 4:
            new_reporter = input("Enter reporter: ")
            for i in ws.iter_rows(values_only=True):
                if i[0] == task_id:
                    i[4] = new_reporter
        elif choice == 5:
            print("Edit made successfully")
            break
        else:
            print("Invalid choice")
    wb.save('kanban_board.xlsx')

def tasks_sorted(project):
    df = pd.read_excel('kanban_board.xlsx', sheet_name=project)
    sorted_df = df.sort_values(by='priority')
    with pd.ExcelWriter('kanban_board.xlsx', engine='openpyxl', mode='w') as writer:
        sorted_df.to_excel(writer, index=False, sheet_name=project)
    display_tasks(project)


def edit_project(project):
    wb = load_workbook('kanban_board.xlsx')
    ws = wb[project]
    while True:
        print("Options\n1.Add Status fields\n2.Delete Project\n3.Go Back")
        choice = input("Enter your choice: ")
        if choice == '1':
            new_field = input("Enter the status field: ")
            for i in range(len(status)):
                print(i,status[i])
            new_position = input("Enter the position after which the status comes up: ")
            status.insert(new_field,new_position)
            print("Status field added successfully")
        elif choice == '2':
            print("To confirm deleting : 'y'\nElse: any number")
            y = input()
            if y == 'y':
                ws.remove(project)
                print("Project deleted successfully")
            else:
                print("Project not deleted")
        elif choice == '3':
            break
        else:
            print("Invalid choice")



def main():
    wb = load_workbook('kanban_board.xlsx')
    while True:
        print("\nKanban Board:")
        print("Main Menu:\n1.New project\n2.Current projects\n3.Exit")
        ch = int(input("Enter your choice: "))
        if ch == 1:
            new_project()
        elif ch == 2:
            print("Current projects")
            display_project()
            project = input("Enter project name to see details: ")
            display_tasks(project)
            while True:
                print("\nOptions:")
                print("1. Add Task")
                print("2. Move Task")
                print("3. Remove Task")
                print("4. Edit Task")
                print("5. Display Task")
                print("6. Edit Project Details")
                print("7. Go to Main Menu")
                choice = input("Enter your choice: ")
                if choice == '1':
                    add_task(project)
                elif choice == '2':
                    move_task(project)
                elif choice == '3':
                    remove_task(project)
                elif choice == '4':
                    edit_task(project)
                elif choice == '5':
                    tasks_sorted(project)
                elif choice == '6':
                    edit_project(project)
                elif choice == '7':
                    print("Going back to main menu")
                    break
                else:
                    print("Invalid choice. Please try again.")
        elif ch == 3:
            print("Exiting program")
            break
        else:
            print("Invalid choice")


if __name__ == "__main__":
    main()
